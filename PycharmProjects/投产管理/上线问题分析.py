# -*- coding: utf-8 -*-
# @Time : 2022-01-18
# @Author : dwd
# @File : 上线问题分析.py

from openpyxl import load_workbook
import win32com.client
import os
import datetime
from collections import Counter
import pandas as pd
from tqdm import tqdm


def replace_excel(folder_path, file_name):
    """
    excel  .xls 后缀 改成 .xlsx 后缀
    folder_path 文件夹路径
    file_name 文件名字 带后缀 比如 aa.xls
    """
    name, suffix = file_name.split('.')
    excel_file_path = os.path.join(folder_path, file_name)

    excel = win32com.client.gencache.EnsureDispatch('Excel.Application')  # 要看MIME手册
    wb = excel.Workbooks.Open(excel_file_path)
    suffix = f".{suffix}x"
    new_file_name = f"{name}{suffix}"
    new_excel_file_path = os.sep.join([folder_path, new_file_name])
    wb.SaveAs(new_excel_file_path, FileFormat=51)
    wb.Close()
    excel.Application.Quit()
    os.remove(excel_file_path)
    return new_excel_file_path
def generate_report():
    wb1 = load_workbook(import_file_path1, data_only=False)  # 上线日志原始版
    wb2 = load_workbook(import_file_path2, data_only=False)  # 上线项目统计表原始版
    # wb3 = load_workbook(export_file_path, data_only=False)  # 上线日志导出版，废除
    wb4 = load_workbook(import_file_path3, data_only=False)  # 获取网盘导出版的数据
    wb5 = load_workbook(renyuan_file_path, data_only=False)  # 获取网盘导出版的数据




    _project_info(wb1, wb2, wb4,wb5)
    _check_apply_time(wb1)  # 生成业务超期提交上线申请报告
    _check_process_confirm_time(wb1)# 生成宗总确认时间报告
    _check_technical_manager_delete_mistakes(wb1) #生成技术经理多填写的报告
    _check_technical_manager_add_mistakes(wb1) #生成技术经理少填写的报告
    approval_time_list(wb1)  # 生成审批时间表格，并将生成内容存到当前目录下

    wb1.save(export_file_path)
    wb1.close()
def _project_info(wb1,wb2,wb4,wb5):
    '''
    加工项目信息
    '''
    sheet3 = wb2['上线项目统计表']
    sheet4 = wb4['上线统计']
    sheet8 = wb1.create_sheet('上线项目统计表', 11)
    sheet8.sheet_state = 'hidden'
    for value in sheet3.iter_rows(min_row=1, max_row=sheet3.max_row, min_col=1, max_col=sheet3.max_column,
                                  values_only=True):
        value = list(value)
        sheet8.append(value)
    sheet9 = wb5['Sheet1']
    sheet10 = wb1.create_sheet('人员信息表', 12)
    sheet10.sheet_state = 'hidden'
    for value in sheet9.iter_rows(min_row=1, max_row=sheet9.max_row, min_col=1, max_col=sheet9.max_column,
                                  values_only=True):
        value = list(value)
        sheet10.append(value)

    terminate_pid = []  # 流程终止的项目编号
    global export_pid
    export_pid = []  # 网盘导出的项目编号
    global log_pid
    log_pid = []  # 上线日志中的项目编号
    repeat = []  # 记录重复的项目编号
    global void_pid
    void_pid = []  # 无效的项目编号


    # 统计网盘填写的上线项目数量
    export_pnum = sheet4.max_row - 2

    # 统计项目上线统计表中项目的项目编号
    for row in sheet3.iter_rows(min_row=2, max_row=sheet3.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value is not None:
                log_pid.append(cell.value)
        # 计算编号重复的次数，并转换为字典
        dict_project_nums = dict(Counter(log_pid))
        # 统计编号重复次数大于1的项目，并将编号记录进repeat列表中
        for key, value in dict_project_nums.items():
            if value > 1 and key not in repeat:
                repeat.append(key)
        log_pid = set(log_pid)
        log_pid = list(log_pid)
    # 统计上线流程强行终止或驳回的项目编号
    for row in sheet3.iter_rows(min_row=2, max_row=sheet3.max_row, min_col=17, max_col=17):
        for cell in row:
            if cell.value == '强行终止' or cell.value == '驳回':
                terminate_pid.append(sheet3.cell(cell.row, 1).value)
    for i in terminate_pid:
        if i not in repeat:
            void_pid.append(i)
    # 统计网盘中的项目编号
    for row in sheet4.iter_rows(min_row=3, max_row=sheet4.max_row, min_col=2, max_col=2):
        for cell in row:
            export_pid.append(cell.value)
def _check_apply_time(wb1):
    '''
        完成检核规则1功能，业务人员未在规定时间发起上线申请 = （I列操作步骤为流程发起 and J列操作时间>规定的最晚提交时间 ）的项目明细，
        写入到一个sheet页中
    '''

    sheet1 = wb1['项目上线日志']
    sheet1.sheet_state = 'hidden'
    sheet2 = wb1.create_sheet('业务人员未在规定时间发起',3)


    apply_row = []
    overtime_row = []
    overtime_pno = []

    # 记录上线流程发起的行数
    for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=9, max_col=9):#记录流程发起的行数
        for cell in row:
            if  cell.value == '流程发起':
                apply_row.append(cell.row)

    apply_time = datetime.datetime.strptime(apply_times, "%Y-%m-%d %H:%M:%S") #将最晚上线时间由str类型转换成datetime类型
    # 在流程发起操作步骤中，查找超过提交时间的行数
    for i in apply_row:
        for row in sheet1.iter_rows(min_row=i, max_row=i, min_col=10, max_col=10):
            for cell in row:
                a = datetime.datetime.strptime(cell.value, "%Y-%m-%d %H:%M:%S")
                if  a > apply_time:
                    overtime_row.append(cell.row)
    #统计超时提交申请的项目编号
    for i in overtime_row:
        for row in sheet1.iter_rows(min_row=i, max_row=i, min_col=2, max_col=2):
            for cell in row:
                overtime_pno.append(cell.value)
    overtime_pno = set(overtime_pno)
    overtime_pno = list(overtime_pno)
    # print('业务人员未按要求发起上线申请的项目编号为{}'.format(overtime_pno))
    # print('生成业务人员未在规定时间发起上线申请报告')
    #补充表头
    sheet2.cell(1, 1).value = '序号'
    sheet2.cell(1, 2).value = '项目编号'
    sheet2.cell(1, 3).value = '项目名称'
    sheet2.cell(1, 4).value = '业务人员'
    sheet2.cell(1, 5).value = '所属部门'
    sheet2.cell(1, 6).value = '提交时间'
    #补充序号
    for index in range(2, len(overtime_pno) + 2):
        sn = 'A' + str(index)
        sheet2[sn] = index - 1
    #补充项目编号
    index = 1
    for pno in overtime_pno:
        index += 1
        sheet2.cell(index, 2).value = pno
    #补充项目名称
    for index in range(2, len(overtime_pno) + 2):
        project_num = 'B' + str(index)
        project_name = 'C' + str(index)
        sheet2[project_name] = '=IFERROR(VLOOKUP(' + project_num + ',项目上线日志!B:C,2,0),"")'
    #补充业务人员
    for index in range(2, len(overtime_pno) + 2):
        project_num = 'B' + str(index)
        project_business = 'D' + str(index)
        sheet2[project_business] = '=IFERROR(VLOOKUP(' + project_num + ',项目上线日志!B:G,6,0),"")'
    # 补充所属部门
    for index in range(2, len(overtime_pno) + 2):
        project_num = 'B' + str(index)
        business_department = 'E' + str(index)
        # sheet2[business_department] = "=IFERROR(VLOOKUP(" + project_num + ",'[上线项目统计表-20220111-0112导出.xlsx]" \
        #                                                                   "上线项目统计表'!A:C,3,0),"")"
        sheet2[business_department] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:C,3,0),"")"
    # 补充提交时间
    for index in range(2, len(overtime_pno) + 2):
        project_num = 'B' + str(index)
        time = 'F' + str(index)
        sheet2[time] = '=IFERROR(VLOOKUP(' + project_num + ',项目上线日志!B:J,9,0),"")'
def _check_process_confirm_time(wb1):
    '''
    生成流程问题报告
    '''
    sheet1 = wb1['项目上线日志']
    sheet6 = wb1.create_sheet('上线日17点后运维领导审核的项目', 5)
    sheet7 = wb1.create_sheet('已上线运维领导未审核项目', 6)

    zongyongtao_confirm_between_12_17 = []  # 17点之前
    zongyongtao_confirm_before_17 = []  # 17点之前
    zongyongtao_confirm_after_17 = []  # 17点之后
    zongyongtao_not_confirm = []  # 未流转
    global not_confirm_not_production_pid
    not_confirm_not_production_pid = []  # 没有上线的项目
    not_confirm_production_pid = []
    # 统计投产日前一天12:00到投产日17:00之间宗总审核的项目数量

    for row in sheet1.iter_rows(min_row=3, max_row=sheet1.max_row, min_col=10, max_col=10):
        for cell in row:
            a = datetime.datetime.strptime(sheet1.cell(cell.row - 1, 10).value, "%Y-%m-%d %H:%M:%S")
            if a < production_time and a > production_times_the_day_before and sheet1.cell(cell.row, 7).value == '宗勇涛':
                zongyongtao_confirm_between_12_17.append(sheet1.cell(cell.row, 2).value) #取到宗总的时间而不是审批的时间
    print('投产日前一天12:00到投产日17:00之间宗总审核的项目数量：{}'.format(len(zongyongtao_confirm_between_12_17 )))

    # 统计17点前流转到宗总的项目

    for row in sheet1.iter_rows(min_row=3, max_row=sheet1.max_row, min_col=10, max_col=10):
        for cell in row:
            a = datetime.datetime.strptime(sheet1.cell(cell.row - 1, 10).value, "%Y-%m-%d %H:%M:%S")
            if a < production_time and sheet1.cell(cell.row, 7).value == '宗勇涛' :
                zongyongtao_confirm_before_17.append(sheet1.cell(cell.row, 2).value)
    # print('17:00之前宗总审核的项目数量：{}'.format(len(zongyongtao_confirm_before_17)))

    # 统计17点后流转到宗总的项目

    for row in sheet1.iter_rows(min_row=3, max_row=sheet1.max_row, min_col=10, max_col=10):
        for cell in row:
            a = datetime.datetime.strptime(sheet1.cell(cell.row - 1, 10).value, "%Y-%m-%d %H:%M:%S")
            if a > production_time and sheet1.cell(cell.row, 7).value == '宗勇涛':
                zongyongtao_confirm_after_17.append(sheet1.cell(cell.row, 2).value)
    # print('17:00之后宗总审核的项目{}'.format(zongyongtao_confirm_after_17))

    # 统计未流转到宗总的项目

    for pid in log_pid:
        if pid not in zongyongtao_confirm_before_17 and pid not in zongyongtao_confirm_after_17:
            zongyongtao_not_confirm.append(pid)
    zongyongtao_not_confirm = set(zongyongtao_not_confirm)
    zongyongtao_not_confirm = list(zongyongtao_not_confirm)
    print('未流转到宗总的项目{}'.format(zongyongtao_not_confirm))
    for i in zongyongtao_not_confirm:
        while True:
            confirm = input('请确认{}实际是否上线,请输入y或n：'.format(i))
            if  confirm == 'n':
                not_confirm_not_production_pid.append(i)
                break
            elif confirm == 'y':
                not_confirm_production_pid.append(i)
                break
            else:
                print('输入不正确，请重新输入')

    # print('未流转到宗总且实际未投产{}'.format(not_confirm_not_production_pid))
    # print('未流转到宗总且实际投产{}'.format(not_confirm_production_pid))
    # print('----------将以上结果生成报告----------')
    # print('生成上线日17点后运维领导审核的项目表')
    # 补充表头
    sheet6.cell(1, 1).value = '序号'
    sheet6.cell(1, 2).value = '项目编号'
    sheet6.cell(1, 3).value = '项目名称'
    sheet6.cell(1, 4).value = '业务'
    sheet6.cell(1, 5).value = '运营'
    sheet6.cell(1, 6).value = '技术负责人'
    sheet6.cell(1, 7).value = '所属室'
    sheet6.cell(1, 8).value = '项目经理'
    sheet6.cell(1, 9).value = '备注'

    # 补充序号
    for index in range(2, len(zongyongtao_confirm_after_17) + 2):
        sn = 'A' + str(index)
        sheet6[sn] = index - 1
    # 补充项目编号
    index = 1
    for pid in zongyongtao_confirm_after_17:
        index += 1
        sheet6.cell(index, 2).value = pid
    # 补充项目名称
    for index in range(2, len(zongyongtao_confirm_after_17) + 2):
        project_num = 'B' + str(index)
        project_name = 'C' + str(index)
        sheet6[project_name] = '=IFERROR(VLOOKUP(' + project_num + ',项目上线日志!B:C,2,0),"")'
    #补充业务负责人
    for index in range(2, len(zongyongtao_confirm_after_17) + 2):
        project_num = 'B' + str(index)
        business_manager = 'D' + str(index)
        sheet6[business_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:F,6,0),"")"
    # 补充运营负责人
    for index in range(2, len(zongyongtao_confirm_after_17) + 2):
        project_num = 'B' + str(index)
        operation_manager = 'E' + str(index)
        sheet6[operation_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:N,14,0),"")"
    #补充技术负责人
    for index in range(2, len(zongyongtao_confirm_after_17) + 2):
        project_num = 'B' + str(index)
        technical_manager = 'F' + str(index)
        sheet6[technical_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:H,8,0),"")"
    # 补充技术负责人所属室
    for index in range(2, len(zongyongtao_confirm_after_17) + 2):
        technical_manage_name = 'F' + str(index)
        technical_manager_group = 'G' + str(index)
        sheet6[technical_manager_group] = "=IFERROR(VLOOKUP(" + technical_manage_name + ",人员信息表!C:I,7,0),"")"

    #补充项目经理
    for index in range(2, len(zongyongtao_confirm_after_17) + 2):
        project_num = 'B' + str(index)
        project_manager = 'H' + str(index)
        sheet6[project_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:G,7,0),"")"

    print('生成已上线运维领导未审核项目')
    # 补充表头
    sheet7.cell(1, 1).value = '序号'
    sheet7.cell(1, 2).value = '项目编号'
    sheet7.cell(1, 3).value = '项目名称'
    sheet7.cell(1, 4).value = '业务'
    sheet7.cell(1, 5).value = '运营'
    sheet7.cell(1, 6).value = '技术负责人'
    sheet7.cell(1, 7).value = '所属室'
    sheet7.cell(1, 8).value = '项目经理'
    sheet7.cell(1, 9).value = '备注'

    # 补充序号
    for index in range(2, len(not_confirm_production_pid) + 2):
        sn = 'A' + str(index)
        sheet7[sn] = index - 1
    # 补充项目编号
    index = 1
    for pid in not_confirm_production_pid:
        index += 1
        sheet7.cell(index, 2).value = pid
    # 补充项目名称
    for index in range(2, len(not_confirm_production_pid) + 2):
        project_num = 'B' + str(index)
        project_name = 'C' + str(index)
        sheet7[project_name] = '=IFERROR(VLOOKUP(' + project_num + ',项目上线日志!B:C,2,0),"")'
    # 补充业务负责人
    for index in range(2, len(not_confirm_production_pid) + 2):
        project_num = 'B' + str(index)
        business_manager = 'D' + str(index)
        sheet7[business_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:F,6,0),"")"
    # 补充运营负责人
    for index in range(2, len(not_confirm_production_pid) + 2):
        project_num = 'B' + str(index)
        operation_manager = 'E' + str(index)
        sheet7[operation_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:N,14,0),"")"
    # 补充技术负责人
    for index in range(2, len(not_confirm_production_pid) + 2):
        project_num = 'B' + str(index)
        technical_manager = 'F' + str(index)
        sheet7[technical_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:H,8,0),"")"
    # 补充技术负责人所属室
    for index in range(2, len(not_confirm_production_pid) + 2):
        technical_manage_name = 'F' + str(index)
        technical_manager_group = 'G' + str(index)
        sheet7[technical_manager_group] = "=IFERROR(VLOOKUP(" + technical_manage_name + ",人员信息表!C:I,7,0),"")"
    # 补充项目经理
    for index in range(2, len(not_confirm_production_pid) + 2):
        project_num = 'B' + str(index)
        project_manager = 'H' + str(index)
        sheet7[project_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:G,7,0),"")"
def _check_technical_manager_delete_mistakes(wb1):
    '''
    生成登记问题报告
    '''
    sheet5 = wb1.create_sheet('技术经理多填写上线登记', 2)
    technical_manager_delete_mistakes = []  # 技术经理多填写上线登记

    #统计网盘没有删除的无效项目编号，网盘中存在的没有删除的流程终止的项目+未流转到宗总且没有上线
    for i in export_pid:
        if i in void_pid or i in not_confirm_not_production_pid:
            technical_manager_delete_mistakes.append(i)
    print('技术经理多填写上线登记项目：{}'.format(technical_manager_delete_mistakes))
    # 补充表头
    sheet5.cell(1, 1).value = '序号'
    sheet5.cell(1, 2).value = '项目编号'
    sheet5.cell(1, 3).value = '项目名称'
    sheet5.cell(1, 4).value = '技术负责人'
    sheet5.cell(1, 5).value = '所属室'

    # 补充序号
    for index in range(2, len(technical_manager_delete_mistakes) + 2):
        sn = 'A' + str(index)
        sheet5[sn] = index - 1
    # 补充项目编号
    index = 1
    for pid in technical_manager_delete_mistakes:
        index += 1
        sheet5.cell(index, 2).value = pid
    # 补充项目名称
    for index in range(2, len(technical_manager_delete_mistakes) + 2):
        project_num = 'B' + str(index)
        project_name = 'C' + str(index)
        sheet5[project_name] = '=IFERROR(VLOOKUP(' + project_num + ',项目上线日志!B:C,2,0),"")'
    #补充技术负责人
    for index in range(2, len(technical_manager_delete_mistakes) + 2):
        project_num = 'B' + str(index)
        technical_manager = 'D' + str(index)
        sheet5[technical_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:H,8,0),"")"
    # 补充技术负责人所属室
    for index in range(2, len(technical_manager_delete_mistakes) + 2):
        technical_manage_name = 'D' + str(index)
        technical_manager_group = 'E' + str(index)
        sheet5[technical_manager_group] = "=IFERROR(VLOOKUP(" + technical_manage_name + ",人员信息表!C:I,7,0),"")"
def _check_technical_manager_add_mistakes(wb1):
    sheet11 = wb1.create_sheet('技术经理少填写上线登记', 1)
    pid = [] #有效上线数量
    technical_manager_add_mistakes = []  # 技术经理多填写上线登记

    for i in log_pid:
        if i not in not_confirm_not_production_pid:
            pid.append(i)
    print('有效上线项目数量:{}'.format(len(pid)))
    for i in pid:
        if i not in export_pid:
            technical_manager_add_mistakes.append(i)
    print('技术经理少填的项目：{}'.format(technical_manager_add_mistakes))
    # 补充表头
    sheet11.cell(1, 1).value = '序号'
    sheet11.cell(1, 2).value = '项目编号'
    sheet11.cell(1, 3).value = '项目名称'
    sheet11.cell(1, 4).value = '技术负责人'
    sheet11.cell(1, 5).value = '所属室'
    # 补充序号
    for index in range(2, len(technical_manager_add_mistakes) + 2):
        sn = 'A' + str(index)
        sheet11[sn] = index - 1
    # 补充项目编号
    index = 1
    for pid in technical_manager_add_mistakes:
        index += 1
        sheet11.cell(index, 2).value = pid
    # 补充项目名称
    for index in range(2, len(technical_manager_add_mistakes) + 2):
        project_num = 'B' + str(index)
        project_name = 'C' + str(index)
        sheet11[project_name] = '=IFERROR(VLOOKUP(' + project_num + ',项目上线日志!B:C,2,0),"")'
    #补充技术负责人
    for index in range(2, len(technical_manager_add_mistakes) + 2):
        project_num = 'B' + str(index)
        technical_manager = 'D' + str(index)
        sheet11[technical_manager] = "=IFERROR(VLOOKUP(" + project_num + ",上线项目统计表!A:H,8,0),"")"
    # 补充技术负责人所属室
    for index in range(2, len(technical_manager_add_mistakes) + 2):
        technical_manage_name = 'D' + str(index)
        technical_manager_group = 'E' + str(index)
        sheet11[technical_manager_group] = "=IFERROR(VLOOKUP(" + technical_manage_name + ",人员信息表!C:I,7,0),"")"
def time_overlap(t):
    #计算两个日期间重叠的时间，返回重叠的小时、分钟、秒
    overlap = (min([x[1] for x in t]) - max([x[0] for x in t]))
    days = overlap.days
    hours, remainder = divmod(overlap.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    hours += days*24
    return [hours,minutes,seconds]
def time(t1,t2):
    #将重叠的时间相加，工作时间为8:30-12:00,13:30-18：00，并转换为小时数，并返回相差的小时数，精确到分钟
    if time_overlap(t1)[0] >= 0:
        a1 = time_overlap(t1)
    else:
        a1 = [0,0,0]
    if time_overlap(t2)[0] >= 0:
        a2 = time_overlap(t2)
    else:
        a2 = [0,0,0]
    a = '{}:{}:{}'.format(a1[0],a1[1],a1[2])
    b = datetime.datetime.strptime(a, "%H:%M:%S") + datetime.timedelta(hours=a2[0],minutes=a2[1])
    h = float('{}'.format(b.hour)) + float('{}'.format(b.minute)) / 60
    return h
def getBetweenDay(begin_date,end_date):
    #计算两个日期间的工作日天数（含调休日），刨除双休日和节假日。
    wb11 = load_workbook(date_match_file_path, data_only=False)  # 工作日历
    sheet11 = wb11['date_match']
    end_day = 0
    begin_day = 0
    for row in sheet11.iter_rows(min_row=2, max_row=sheet11.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value == begin_date :
                begin_day = (sheet11.cell(cell.row,4).value)
            if cell.value == end_date:
                end_day = (sheet11.cell(cell.row,4).value)
    return end_day - begin_day
def _check_workday(begin_date):
    # 计算日期是否为工作日，并返回工作日，双休日，法定节假日，调休日，目前列表中储存的为20210930-20220630
    wb11 = load_workbook(date_match_file_path, data_only=False)  # 工作日历
    sheet12 = wb11['date_match']
    for row in sheet12.iter_rows(min_row=2, max_row=sheet12.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value == begin_date:
                return sheet12.cell(cell.row,5).value
def _check_begin_time(begin_time):
    #判断开始时间是否在双休或者是节假日中，如果在节假日中，顺延到下一个工作日的8:30
    a = begin_time.date()
    i = 0
    while True:
        if _check_workday(int(a.strftime("%Y%m%d"))) == '工作日' or _check_workday(int(a.strftime("%Y%m%d"))) == '调休日':
            break
        else:
            a += datetime.timedelta(days=1)
            i += 1
    begin_time = begin_time + datetime.timedelta(days=i)
    if i != 0:
        begin_time = begin_time.replace(hour=8, minute=30, second=0)
    return begin_time
def _check_end_time(end_time):
    # 判断结束时间是否在双休或者是节假日中，如果在节假日中，顺延到上一个工作日的18:00
    a = end_time.date()
    i = 0
    while True:
        if _check_workday(int(a.strftime("%Y%m%d"))) == '工作日' or _check_workday(int(a.strftime("%Y%m%d"))) == '调休日':
            break
        else:
            a -= datetime.timedelta(days=1)
            i += 1
    end_time = end_time + datetime.timedelta(days=-i)
    if i != 0:
        end_time = end_time.replace(hour=18,minute=0,second=0)
    return end_time
def sum(list):
    #列表中的元素求和
    sum = 0
    for i in list:
        sum += i
    return sum
def count_0_4(list):
    count = 0
    for i in list:
        if i <4 and i >= 0:
            count += 1
    return count
def count_4_8(list):
    count = 0
    for i in list:
        if i <8 and i >= 4:
            count += 1
    return count
def count_8_50(list):
    count = 0
    for i in list:
        if i <50 and i >= 8:
            count += 1
    return count
def count_50(list):
    count = 0
    for i in list:
        if  i >= 50:
            count += 1
    return count
def approval_time(pno):
    #核心函数，生成一个项目的各个阶段的处理时间，返回一个处理时间的列表和处理时间汇总列表的元祖集

    process_type = ['提交上线申请','开发负责人审核','系统负责人上传文档','开发负责人提供上线文档','运营代表审批','项目经理审核上线申请表','业务部领导审批','业务线领导审批','开发部领导审批','运维部领导审批','执行人执行','验证人验证']
    提交上线申请 = []
    开发负责人审核 = []
    系统负责人上传文档 = []
    开发负责人提供上线文档 = []
    运营代表审批 = []
    项目经理审核上线申请表 = []
    业务部领导审批 = []
    业务线领导审批 = []
    开发部领导审批 = []
    运维部领导审批 = []
    执行人执行 = []
    验证人验证 = []
    wb1 = load_workbook(import_file_path1, data_only=False)  # 上线日志原始版
    sheet1 = wb1['项目上线日志']
    for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=2, max_col=2):
        for cell in row:
            for i in process_type:
                if i == sheet1.cell(cell.row,9).value:
                    if cell.value == pno:
                        end_time = datetime.datetime.strptime(sheet1.cell(cell.row, 10).value, "%Y-%m-%d %H:%M:%S")
                        begin_time = datetime.datetime.strptime(sheet1.cell(cell.row - 1, 10).value, "%Y-%m-%d %H:%M:%S")
                        if end_time.date() == begin_time.date() and _check_workday(int(begin_time.date().strftime("%Y%m%d"))) == '工作日':#需要补充是否日期在节假日列表中
                            # print('开始与结束在一天')
                            a = begin_time.replace(hour=8, minute=30, second=00)
                            b = end_time.replace(hour=12, minute=00, second=00)
                            c = begin_time.replace(hour=13, minute=30, second=00)
                            d = end_time.replace(hour=18, minute=00, second=00)
                            t1 = [(a, b),(begin_time, end_time)]
                            t2 = [(c, d), (begin_time, end_time)]
                            processing_time_day = round(time(t1, t2),2)
                            # print('{}耗时{}'.format(i,processing_time_day))
                            if  i == process_type[0]:
                                提交上线申请.append(processing_time_day)
                            elif i == process_type[1]:
                                开发负责人审核.append(processing_time_day)
                            elif i == process_type[2]:
                                系统负责人上传文档.append(processing_time_day)
                            elif i == process_type[3]:
                                开发负责人提供上线文档.append(processing_time_day)
                            elif i == process_type[4]:
                                运营代表审批.append(processing_time_day)
                            elif i == process_type[5]:
                                项目经理审核上线申请表.append(processing_time_day)
                            elif i == process_type[6]:
                                业务部领导审批.append(processing_time_day)
                            elif i == process_type[7]:
                                业务线领导审批.append(processing_time_day)
                            elif i == process_type[8]:
                                开发部领导审批.append(processing_time_day)
                            elif i == process_type[9]:
                                运维部领导审批.append(processing_time_day)
                            elif i == process_type[10]:
                                执行人执行.append(processing_time_day)
                            elif i == process_type[11]:
                                验证人验证.append(processing_time_day)
                        else:
                            #判断开始时间和结束时间是否在工作日
                            if _check_workday(int(begin_time.date().strftime("%Y%m%d"))) == '工作日' and _check_workday(int(end_time.date().strftime("%Y%m%d"))) == '工作日':
                                # print('开始与结束不在一天，但开始和结束时间不再节假日中')
                                e = begin_time.replace(hour=8, minute=30, second=00)
                                f = begin_time.replace(hour=12, minute=00, second=00)
                                g = begin_time.replace(hour=13, minute=30, second=00)
                                h = begin_time.replace(hour=18, minute=00, second=00)
                                t1 = [(e, f), (begin_time, h)]
                                t2 = [(g, h), (begin_time, h)]
                                time1 = time(t1, t2)
                                n = end_time.replace(hour=8, minute=30, second=00)
                                j = end_time.replace(hour=12, minute=00, second=00)
                                k = end_time.replace(hour=13, minute=30, second=00)
                                l = end_time.replace(hour=18, minute=00, second=00)
                                m = end_time.replace(hour=0, minute=00, second=00)
                                t1 = [(n, j), (m, end_time)]
                                t2 = [(k, l), (m, end_time)]
                                time2 = time(t1, t2)
                                day = getBetweenDay(int(begin_time.date().strftime("%Y%m%d")),int(end_time.date().strftime("%Y%m%d")))
                                processing_time_days = round(time1 + time2 + ((day + 1 - 2) * 8),2)
                                # print('{}耗时{}'.format(i,processing_time_days))
                                if  i == process_type[0]:
                                    提交上线申请.append(processing_time_days)
                                elif i == process_type[1]:
                                    开发负责人审核.append(processing_time_days)
                                elif i == process_type[2]:
                                    系统负责人上传文档.append(processing_time_days)
                                elif i == process_type[3]:
                                    开发负责人提供上线文档.append(processing_time_days)
                                elif i == process_type[4]:
                                    运营代表审批.append(processing_time_days)
                                elif i == process_type[5]:
                                    项目经理审核上线申请表.append(processing_time_days)
                                elif i == process_type[6]:
                                    业务部领导审批.append(processing_time_days)
                                elif i == process_type[7]:
                                    业务线领导审批.append(processing_time_days)
                                elif i == process_type[8]:
                                    开发部领导审批.append(processing_time_days)
                                elif i == process_type[9]:
                                    运维部领导审批.append(processing_time_days)
                                elif i == process_type[10]:
                                    执行人执行.append(processing_time_days)
                                elif i == process_type[11]:
                                    验证人验证.append(processing_time_days)
                                else:
                                    pass
                            else:
                                # print('开始与结束不在一天，但开始和结束时间在节假日中')
                                # print(begin_time)
                                # print(end_time)
                                begin_time = _check_begin_time(begin_time)
                                end_time = _check_end_time(end_time)
                                # print(begin_time)
                                # print(end_time)
                                e = begin_time.replace(hour=8, minute=30, second=00)
                                f = begin_time.replace(hour=12, minute=00, second=00)
                                g = begin_time.replace(hour=13, minute=30, second=00)
                                h = begin_time.replace(hour=18, minute=00, second=00)
                                t1 = [(e, f), (begin_time, h)]
                                t2 = [(g, h), (begin_time, h)]
                                time1 = time(t1, t2)
                                n = end_time.replace(hour=8, minute=30, second=00)
                                j = end_time.replace(hour=12, minute=00, second=00)
                                k = end_time.replace(hour=13, minute=30, second=00)
                                l = end_time.replace(hour=18, minute=00, second=00)
                                m = end_time.replace(hour=0, minute=00, second=00)
                                t1 = [(n, j), (m, end_time)]
                                t2 = [(k, l), (m, end_time)]
                                time2 = time(t1, t2)
                                day = getBetweenDay(int(begin_time.date().strftime("%Y%m%d")),
                                                    int(end_time.date().strftime("%Y%m%d")))
                                processing_time_days = round(time1 + time2 + ((day + 1 - 2) * 8),2)
                                # print('{}耗时{}'.format(i, processing_time_days))
                                if i == process_type[0]:
                                    提交上线申请.append(processing_time_days)
                                elif i == process_type[1]:
                                    开发负责人审核.append(processing_time_days)
                                elif i == process_type[2]:
                                    系统负责人上传文档.append(processing_time_days)
                                elif i == process_type[3]:
                                    开发负责人提供上线文档.append(processing_time_days)
                                elif i == process_type[4]:
                                    运营代表审批.append(processing_time_days)
                                elif i == process_type[5]:
                                    项目经理审核上线申请表.append(processing_time_days)
                                elif i == process_type[6]:
                                    业务部领导审批.append(processing_time_days)
                                elif i == process_type[7]:
                                    业务线领导审批.append(processing_time_days)
                                elif i == process_type[8]:
                                    开发部领导审批.append(processing_time_days)
                                elif i == process_type[9]:
                                    运维部领导审批.append(processing_time_days)
                                elif i == process_type[10]:
                                    执行人执行.append(processing_time_days)
                                elif i == process_type[11]:
                                    验证人验证.append(processing_time_days)
                                else:
                                    pass
    # print('{}用时为{}小时'.format(process_type[0],sum(提交上线申请)))
    # print('{}用时为{}小时'.format(process_type[1], sum(开发负责人审核)))
    # print('{}用时为{}小时'.format(process_type[2], sum(系统负责人上传文档)))
    # print('{}用时为{}小时'.format(process_type[3], sum(开发负责人提供上线文档)))
    # print('{}用时为{}小时'.format(process_type[4], sum(运营代表审批)))
    # print('{}用时为{}小时'.format(process_type[5], sum(项目经理审核上线申请表)))
    # print('{}用时为{}小时'.format(process_type[6], sum(业务部领导审批)))
    # print('{}用时为{}小时'.format(process_type[7], sum(业务线领导审批)))
    # print('{}用时为{}小时'.format(process_type[8], sum(开发部领导审批)))
    # print('{}用时为{}小时'.format(process_type[9], sum(运维部领导审批)))
    # print('{}用时为{}小时'.format(process_type[10], sum(执行人执行)))
    # print('{}用时为{}小时'.format(process_type[10], sum(验证人验证)))
    return (
    [提交上线申请, 开发负责人审核, 系统负责人上传文档, 开发负责人提供上线文档, 运营代表审批, 项目经理审核上线申请表, 业务部领导审批, 业务线领导审批, 开发部领导审批, 运维部领导审批, 执行人执行, 验证人验证],
    [sum(提交上线申请), sum(开发负责人审核), sum(系统负责人上传文档), sum(开发负责人提供上线文档), sum(运营代表审批), sum(项目经理审核上线申请表), sum(业务部领导审批),sum(业务线领导审批), sum(开发部领导审批), sum(运维部领导审批), sum(执行人执行), sum(验证人验证)],
    [len(提交上线申请), len(开发负责人审核), len(系统负责人上传文档), len(开发负责人提供上线文档), len(运营代表审批), len(项目经理审核上线申请表), len(业务部领导审批),len(业务线领导审批), len(开发部领导审批), len(运维部领导审批), len(执行人执行), len(验证人验证)],
    [count_0_4(提交上线申请), count_0_4(开发负责人审核), count_0_4(系统负责人上传文档), count_0_4(开发负责人提供上线文档), count_0_4(运营代表审批),
     count_0_4(项目经理审核上线申请表), count_0_4(业务部领导审批),
     count_0_4(业务线领导审批), count_0_4(开发部领导审批), count_0_4(运维部领导审批), count_0_4(执行人执行), count_0_4(验证人验证)],
    [count_4_8(提交上线申请), count_4_8(开发负责人审核), count_4_8(系统负责人上传文档), count_4_8(开发负责人提供上线文档), count_4_8(运营代表审批),
     count_4_8(项目经理审核上线申请表), count_4_8(业务部领导审批),
     count_4_8(业务线领导审批), count_4_8(开发部领导审批), count_4_8(运维部领导审批), count_4_8(执行人执行), count_4_8(验证人验证)],
    [count_8_50(提交上线申请), count_8_50(开发负责人审核), count_8_50(系统负责人上传文档), count_8_50(开发负责人提供上线文档), count_8_50(运营代表审批),
     count_8_50(项目经理审核上线申请表), count_8_50(业务部领导审批),
     count_8_50(业务线领导审批), count_8_50(开发部领导审批), count_8_50(运维部领导审批), count_8_50(执行人执行), count_8_50(验证人验证)],
    [count_50(提交上线申请), count_50(开发负责人审核), count_50(系统负责人上传文档), count_50(开发负责人提供上线文档), count_50(运营代表审批),
     count_50(项目经理审核上线申请表), count_50(业务部领导审批),
     count_50(业务线领导审批), count_50(开发部领导审批), count_50(运维部领导审批), count_50(执行人执行), count_50(验证人验证)]
    )
def approval_time_list(wb1):

    process_type = ['提交上线申请', '开发负责人审核', '系统负责人上传文档', '开发负责人提供上线文档', '运营代表审批', '项目经理审核上线申请表', '业务部领导审批', '业务线领导审批',
                    '开发部领导审批', '运维部领导审批', '执行人执行', '验证人验证']
    data = pd.DataFrame()
    pnos = ['P20211102000001', 'P20211008000002', 'P20211012000003']
    # print(log_pid)
    for pno in tqdm(log_pid):
        process_list = pd.DataFrame(
            {'项目编号':pno,'处理类型': process_type, '处理时长明细': approval_time(pno)[0],'处理时长汇总': approval_time(pno)[1],'总条数':approval_time(pno)[2],'0-4小时个数':approval_time(pno)[3],'4-8小时个数':approval_time(pno)[4],'8-50小时个数':approval_time(pno)[5],'50小时个数':approval_time(pno)[6]})
        data = data.append(process_list, ignore_index=True)  # ignore_index=True 连续序号]

    writer = pd.ExcelWriter(export_file_path, engine='openpyxl')
    writer.book = wb1
    data.to_excel(writer, sheet_name="审批环节时间")
    #根据表格统计数据
    #补充统计表头
    sheet12 = wb1['审批环节时间']
    sheet12.cell(1, 12).value = '审批时长'
    sheet12.cell(1, 13).value = '条数'
    sheet12.cell(1, 14).value = '占比'
    sheet12.cell(2, 12).value = '0-4小时'
    sheet12.cell(3, 12).value = '4-8小时'
    sheet12.cell(4, 12).value = '8-50小时'
    sheet12.cell(5, 12).value = '>50小时'
    sheet12.cell(6, 12).value = '合计'
    sheet12.cell(2, 13).value = '=sum(G:G)'
    sheet12.cell(3, 13).value = '=sum(H:H)'
    sheet12.cell(4, 13).value = '=sum(I:I)'
    sheet12.cell(5, 13).value = '=sum(J:J)'
    sheet12.cell(6, 13).value = '=sum(F:F)'
    #补充占比
    for index in range(2, 7):
        process_time_num = 'N' + str(index)
        sheet12[process_time_num] = '=M' + str(index) + '/M6'
        sheet12[process_time_num].number_format = '0.0%'
    sheet12.cell(1, 16).value = '审批时长超过8小时'
    sheet12.cell(1, 17).value = '条数'
    sheet12.cell(1, 18).value = '占比'
    #补充审批类型
    i = 0
    for index in range(2, len(process_type) + 2):
        process_type_name = 'P' + str(index)
        sheet12[process_type_name] = process_type[i]
        i += 1
    #补充审批时长条数
    for index in range(2, len(process_type) + 2):
        process_type_num = 'Q' + str(index)
        sheet12[process_type_num] = '=SUMIFS(I:I,C:C,P' + str(index)+')+SUMIFS(J:J,C:C,P'+ str(index)+')'
        sheet12.cell(14, 17).value = '=SUM(Q2:Q13)'
    #补充审批时长条数占比
    for index in range(2, len(process_type) + 2):
        process_type_num = 'R' + str(index)
        sheet12[process_type_num] = '=Q' + str(index) + '/Q14'
        sheet12[process_type_num].number_format = '0.0%'

if __name__ == '__main__':
    import_file_path1 = r'D:\Users\tc\PycharmProjects\投产管理\0120上线问题分析\项目上线日志-20220120-0121导出.xlsx'#input('请输入原始上线日志：')  # 参考格式D:\Users\tc\PycharmProjects\投产管理\上线问题分析\项目上线日志-20220111-0112导出.xlsx
    import_file_path2 = r'D:\Users\tc\PycharmProjects\投产管理\0120上线问题分析\上线项目统计表-20220120-20220121导出.xlsx'#input('请输入原始上线项目统计表：')  # 参考格式D:\Users\tc\PycharmProjects\投产管理\上线问题分析\上线项目统计表-20220111-0112导出.xlsx
    import_file_path3 = r'D:\Users\tc\PycharmProjects\投产管理\0120上线问题分析\上线日投产项目信息统计-20220120-网盘导出版.xlsx'#input('请输入网盘导出版上线信息统计表：')  # 参考格式D:\Users\tc\PycharmProjects\投产管理\上线日投产项目信息统计-网盘导出版.xlsx
    renyuan_file_path = r'D:\Users\tc\PycharmProjects\投产管理\上线问题分析\20211220-2021年未休年假统计-软件开发中心.xlsx'#input('请输入人员基础信息：')  # 参考格式D:\Users\tc\PycharmProjects\投产管理\上线问题分析\20211220-2021年未休年假统计-软件开发中心.xlsx
    export_file_path = r'D:\Users\tc\PycharmProjects\投产管理\0120上线问题分析\项目上线日志-20220120-分析结果.xlsx'#input('请输入输出地址：')  # 参考格式D:\Users\tc\PycharmProjects\投产管理\上线问题分析\项目上线日志-20220113-分析结果.xlsx
    date_match_file_path = r'D:\Users\tc\PycharmProjects\投产管理\date_match.xlsx'
    apply_times = '2022-1-15  12:00:00'#input('请输入最晚业务提交上线时间（格式：2022-1-6  12:00:00）：')
    production_times = '2022-1-20  17:00:00'#input('请输入投产时间（格式：2022-1-11  17:00:00）:')
    production_time = datetime.datetime.strptime(production_times, "%Y-%m-%d %H:%M:%S")  # 将最晚流转宗总时间由str类型转换成datetime类型
    production_times_the_day_before = production_time+datetime.timedelta(days=-1)+datetime.timedelta(hours=-5)#投产前一天12点
    #将输入的表格统一转换成xlsx格式
    file_path_list = [import_file_path1,import_file_path2,import_file_path3,renyuan_file_path]
    file_path_list1 = []

    for i in file_path_list:
        file = os.path.splitext(i)
        filename, type = file
        filename = os.path.basename(i)
        folder_path = os.path.dirname(i)
        if type == '.xls':
            replace_excel(folder_path, filename)
            portion = os.path.splitext(filename)
            new_name = portion[0] + '.xlsx'
            a = '{}\{}'.format(folder_path, new_name)
            file_path_list1.append(a)
        else:
            a = i
            file_path_list1.append(a)
    import_file_path1 = file_path_list1[0]
    import_file_path2 = file_path_list1[1]
    import_file_path3 = file_path_list1[2]
    renyuan_file_path = file_path_list1[3]

    generate_report()


