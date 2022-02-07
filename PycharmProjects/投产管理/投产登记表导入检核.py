# -*- coding: utf-8 -*-
# @Time : 2022-01-12
# @Author : dwd
# @File : 投产登记表导入检核.py

from openpyxl import load_workbook
from collections import Counter
from openpyxl.styles import PatternFill
from win32com.client import Dispatch

'''
检核规则：
1、检核上线统计sheet页中，项目编号（B列）是否重复值，如有重复值，删除整行内容，并输出删除日志，日志包括项目编号及项目名称。
2、如果技术经理所在室组（H列）为空，则将这行内容挪至增加和删除记录sheet页
3、检核网盘导入版与网盘导出版的差异，导出版项目数量=导入版项目数量-取消上线的项目+业务提交申请晚的项目
'''

# 用于读取公式的值
def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()
def check_duplicate_value():
    '''
    完成检核规则1功能，检核上线统计sheet页中，项目编号（B列）是否重复值，如有重复值，删除整行内容，并输出删除日志，日志包括项目编号及项目名称。
    '''
    # 获取原始表中的项目编号并去重
    wb = load_workbook(file_path)
    sheet1 = wb.get_sheet_by_name('上线统计')
    project_nums = []  # 记录项目编号
    global repeat  # 变为全局变量
    repeat = []  # 记录删除的重复的项目编号

    # 生成一个项目编号列表
    for cell in sheet1['B']:
        if cell.value is not None:
            project_nums.append(cell.value)
        # 计算编号重复的次数，并转换为字典
        dict_project_nums = dict(Counter(project_nums))
        # 统计编号重复次数大于1的项目，并将编号记录进repeat列表中
        for key, value in dict_project_nums.items():
            if value > 1 and key not in repeat:
                repeat.append(key)
        # 返回删除的编号对应的行数
        if cell.value in repeat:
            sheet1.delete_rows(cell.row)
        # 由于删除重复项目后，公式顺序不变，所以更新公式技术经理所在组信息
        for index in range(3, sheet1.max_row + 1):
            manager_name = 'G' + str(index)
            manager_group = 'H' + str(index)
            sheet1[manager_group] = '=IFERROR(VLOOKUP(' + manager_name + ',电话本!B:C,2,0),"")'

    print('重复的项目编号是{},已删除'.format(repeat))
    wb.save('上线日投产项目信息统计-网盘导入版.xlsx')
def delete_null_value():
    # 由于删除重复项目后，存在空行需要数据处理，删除空行
    wb = load_workbook(import_file_path, data_only=False)
    sheet1 = wb.get_sheet_by_name('上线统计')
    null_value = []
    # 删除空白行
    for row in sheet1.iter_rows(min_row=3, max_row=sheet1.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value is None:
                null_value.append(cell.row)
    null_value.sort(reverse=True)
    for row in null_value:
        sheet1.delete_rows(row)
    print('删除空行后，有效行数为{}'.format(sheet1.max_row))
    wb.save('上线日投产项目信息统计-网盘导入版.xlsx')
def check_Technical_Manager():
    '''
    完成检核规则2 功能，如果技术经理所在室组（H列）为空，则将这行内容挪至增加和删除记录sheet页
    '''
    just_open(import_file_path)
    wb = load_workbook(import_file_path, data_only=True)
    sheet1 = wb.get_sheet_by_name('上线统计')
    sheet2 = wb.get_sheet_by_name('增加和删除记录')
    null_value = []
    null_pno = []
    null_pname = []

    # 获取H列为空的行，并将b列项目编号写入到一个列表中
    # 范围从第3行到有值得最后一行，第8列
    for row in sheet1.iter_rows(min_row=3, max_row=sheet1.max_row, min_col=8, max_col=8):
        for cell in row:
            if cell.value is None:
                null_value.append(cell.row)
    for row in sheet1.iter_rows(min_row=3, max_row=sheet1.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.row in null_value:
                null_pno.append(cell.value)
    for row in sheet1.iter_rows(min_row=3, max_row=sheet1.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.row in null_value:
                null_pname.append(cell.value)
    # 将技术经理为空的项目编号写入到添加及删除记录中
    index = 1
    for pno in null_pno:
        index += 1
        sheet2.cell(index, 1).value = pno
    # 填充项目名称
    index = 1
    for pname in null_pname:
        index += 1
        sheet2.cell(index, 2).value = pname
    # 填充项目类型
    for index in range(2, len(null_pno) + 3 - len(repeat)):
        change_name = 'C' + str(index)
        sheet2[change_name] = '删除'
    # 由于去公式值，原有技术经理所在组的公式消失，故重新更新公式技术经理所在组信息
    for index in range(3, sheet1.max_row + 1):
        manager_name = 'G' + str(index)
        manager_group = 'H' + str(index)
        sheet1[manager_group] = '=IFERROR(VLOOKUP(' + manager_name + ',电话本!B:C,2,0),"")'
    # 由于重复编号，更新序号
    for index in range(3, sheet1.max_row + 1):
        sn = 'A' + str(index)
        sheet1[sn] = index - 2
    print('无技术经理所在组的项目编号为{}，这些项目没有删除，需要确认后手动删除'.format(null_pno))
    wb.save('上线日投产项目信息统计-网盘导入版.xlsx')
def check_project():
    '''
    完成检核规则3功能，检核网盘导入版与网盘导出版的差异，导出版项目数量=导入版项目数量-取消上线的项目+业务提交申请晚的项目
    '''
    wb1 = load_workbook(import_file_path, data_only=False)  # 网盘导入版
    sheet1 = wb1.get_sheet_by_name('上线统计')
    sheet2 = wb1.get_sheet_by_name('增加和删除记录')
    wb2 = load_workbook(export_file_path, data_only=False)  # 网盘导出版
    sheet3 = wb2.get_sheet_by_name('上线统计')
    sheet4 = wb2.get_sheet_by_name('增加和删除记录')
    delete_pno = []
    add_pno = []
    null_value = []
    # 删除空白行
    for row in sheet3.iter_rows(min_row=3, max_row=sheet3.max_row, min_col=1, max_col=1):
        for cell in row:
            if cell.value is None:
                null_value.append(cell.row)
    null_value.sort(reverse=True)
    for row in null_value:
        sheet3.delete_rows(row)

    # 统计导入版项目数量
    import_pnum = sheet1.max_row - 2
    # 统计导出版项目数量
    export_pnum = sheet3.max_row - 2
    # 统计删除的项目数量
    for row in sheet4.iter_rows(min_row=2, max_row=sheet4.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value == '删除':
                delete_pno.append(cell.value)
    # 统计增加的项目数量
    for row in sheet4.iter_rows(min_row=2, max_row=sheet4.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value == '增加':
                add_pno.append(cell.value)

    print('导入版的项目数量{}'.format(import_pnum))
    print('导出版的项目数量{}'.format(export_pnum))
    print('删除的项目数量{}'.format(len(delete_pno)))
    print('增加的项目数量{}'.format(len(add_pno)))

    if export_pnum == import_pnum - len(delete_pno) + len(add_pno):
        print('恭喜你，检核结果无误')
    else:
        print('检核结果有误')


file_path = r'D:\Users\tc\PycharmProjects\投产管理\上线日投产项目信息统计-20220111.xlsx'
import_file_path = input("请输入网盘导入版的文件路径：")#（参考格式：D:\Users\tc\PycharmProjects\投产管理\0120\上线日投产项目信息统计-20220120-网盘导入版.xlsx）
export_file_path = input('请输入网盘导出版的文件路径：')#（参考格式：D:\Users\tc\PycharmProjects\投产管理\0120\上线日投产项目信息统计-20220120-网盘导出版.xlsx）
#

# check_duplicate_value()
# delete_null_value()
# check_Technical_Manager()
check_project()




#####################################################################
# fill_white() #行格式问题
# def fill_white():
#     # 设置删除后的单元格格式，填充为白色
#     wb = load_workbook(new_file_path, data_only=False)
#     sheet1 = wb.get_sheet_by_name('上线统计')
#     fille = PatternFill('solid', fgColor='FFFFFF')
#
#     for row in sheet1.iter_rows(min_row=3, max_row=5, min_col=1,max_col=20):
#         for cell in row:
#             cell.fill = fille
#     wb.save('new.xlsx')
